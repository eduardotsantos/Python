from flask import Blueprint, render_template, redirect, url_for, flash, request, send_file
from flask_login import login_required, current_user
from models import db, Project, User, Expense, Resource, Milestone, Timesheet, ProjectCall
from services.tenant_utils import tenant_required, ensure_tenant_access, get_current_tenant_id
from datetime import datetime
from io import BytesIO
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

projects_bp = Blueprint('projects', __name__)


def get_tenant_filter():
    """Get tenant filter for queries."""
    tenant_id = get_current_tenant_id()
    if tenant_id:
        return {'tenant_id': tenant_id}
    return {}


@projects_bp.route('/')
@login_required
@tenant_required
def dashboard():
    return redirect(url_for('projects.list_projects'))


@projects_bp.route('/projects')
@login_required
@tenant_required
def list_projects():
    status_filter = request.args.get('status', '')
    search = request.args.get('search', '')

    tenant_id = get_current_tenant_id()
    query = Project.query
    if tenant_id:
        query = query.filter_by(tenant_id=tenant_id)

    if status_filter:
        query = query.filter_by(status=status_filter)
    if search:
        query = query.filter(
            db.or_(
                Project.title.ilike(f'%{search}%'),
                Project.code.ilike(f'%{search}%'),
                Project.description.ilike(f'%{search}%')
            )
        )

    projects = query.order_by(Project.created_at.desc()).all()

    # Stats for current tenant
    base_query = Project.query
    if tenant_id:
        base_query = base_query.filter_by(tenant_id=tenant_id)

    total = base_query.count()
    in_progress = base_query.filter_by(status='Em Andamento').count()
    completed = base_query.filter_by(status='Concluído').count()
    planning = base_query.filter_by(status='Planejamento').count()

    # Budget stats for tenant
    if tenant_id:
        total_budget = db.session.query(db.func.sum(Project.budget)).filter(Project.tenant_id == tenant_id).scalar() or 0
        total_expenses = db.session.query(db.func.sum(Expense.amount)).filter(
            Expense.tenant_id == tenant_id,
            Expense.status != 'Rejeitada'
        ).scalar() or 0
    else:
        total_budget = db.session.query(db.func.sum(Project.budget)).scalar() or 0
        total_expenses = db.session.query(db.func.sum(Expense.amount)).filter(Expense.status != 'Rejeitada').scalar() or 0

    return render_template('projects/list.html',
                           projects=projects,
                           total=total,
                           in_progress=in_progress,
                           completed=completed,
                           planning=planning,
                           total_budget=total_budget,
                           total_expenses=total_expenses,
                           status_filter=status_filter,
                           search=search)


@projects_bp.route('/projects/new', methods=['GET', 'POST'])
@login_required
@tenant_required
def create_project():
    tenant_id = get_current_tenant_id()

    # Check tenant project limit
    if tenant_id and current_user.tenant and not current_user.tenant.can_add_project():
        flash(f'Limite de projetos atingido ({current_user.tenant.max_projects}). Entre em contato com o suporte.', 'warning')
        return redirect(url_for('projects.list_projects'))

    if request.method == 'POST':
        code = request.form.get('code', '').strip()
        title = request.form.get('title', '').strip()
        description = request.form.get('description', '').strip()
        status = request.form.get('status', 'Planejamento')
        category = request.form.get('category', '')
        start_date_str = request.form.get('start_date', '')
        end_date_str = request.form.get('end_date', '')
        budget = request.form.get('budget', 0)
        funding_source = request.form.get('funding_source', '')

        users = User.query.filter_by(tenant_id=tenant_id).all() if tenant_id else User.query.all()

        if not all([code, title]):
            flash('Código e título são obrigatórios.', 'danger')
            return render_template('projects/form.html', project=None, users=users)

        # Check code uniqueness within tenant
        existing = Project.query.filter_by(tenant_id=tenant_id, code=code).first() if tenant_id else Project.query.filter_by(code=code).first()
        if existing:
            flash('Código do projeto já existe.', 'danger')
            return render_template('projects/form.html', project=None, users=users)

        project = Project(
            tenant_id=tenant_id,
            code=code,
            title=title,
            description=description,
            status=status,
            category=category,
            budget=float(budget) if budget else 0,
            funding_source=funding_source,
            responsible_id=current_user.id
        )

        if start_date_str:
            project.start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        if end_date_str:
            project.end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()

        db.session.add(project)
        db.session.commit()
        flash('Projeto criado com sucesso!', 'success')
        return redirect(url_for('projects.view_project', project_id=project.id))

    users = User.query.filter_by(tenant_id=tenant_id).all() if tenant_id else User.query.all()
    return render_template('projects/form.html', project=None, users=users)


@projects_bp.route('/projects/<int:project_id>')
@login_required
@tenant_required
def view_project(project_id):
    project = Project.query.get_or_404(project_id)
    ensure_tenant_access(project)

    total_expenses = sum(e.amount for e in project.expenses if e.status != 'Rejeitada')
    total_hours = sum(t.hours for t in project.timesheets)
    resource_count = len(project.resources)
    milestone_progress = 0
    if project.milestones:
        milestone_progress = sum(m.progress for m in project.milestones) // len(project.milestones)

    return render_template('projects/view.html',
                           project=project,
                           total_expenses=total_expenses,
                           total_hours=total_hours,
                           resource_count=resource_count,
                           milestone_progress=milestone_progress)


@projects_bp.route('/projects/<int:project_id>/edit', methods=['GET', 'POST'])
@login_required
@tenant_required
def edit_project(project_id):
    project = Project.query.get_or_404(project_id)
    ensure_tenant_access(project)

    tenant_id = get_current_tenant_id()

    if request.method == 'POST':
        new_code = request.form.get('code', '').strip()

        # Check code uniqueness if changed
        if new_code != project.code:
            existing = Project.query.filter_by(tenant_id=tenant_id, code=new_code).first() if tenant_id else Project.query.filter_by(code=new_code).first()
            if existing:
                flash('Código do projeto já existe.', 'danger')
                users = User.query.filter_by(tenant_id=tenant_id).all() if tenant_id else User.query.all()
                return render_template('projects/form.html', project=project, users=users)

        project.code = new_code
        project.title = request.form.get('title', '').strip()
        project.description = request.form.get('description', '').strip()
        project.status = request.form.get('status', 'Planejamento')
        project.category = request.form.get('category', '')
        project.budget = float(request.form.get('budget', 0) or 0)
        project.funding_source = request.form.get('funding_source', '')
        project.responsible_id = request.form.get('responsible_id') or current_user.id

        start_date_str = request.form.get('start_date', '')
        end_date_str = request.form.get('end_date', '')
        if start_date_str:
            project.start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        if end_date_str:
            project.end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()

        db.session.commit()
        flash('Projeto atualizado com sucesso!', 'success')
        return redirect(url_for('projects.view_project', project_id=project.id))

    users = User.query.filter_by(tenant_id=tenant_id).all() if tenant_id else User.query.all()
    return render_template('projects/form.html', project=project, users=users)


@projects_bp.route('/projects/<int:project_id>/delete', methods=['POST'])
@login_required
@tenant_required
def delete_project(project_id):
    project = Project.query.get_or_404(project_id)
    ensure_tenant_access(project)

    db.session.delete(project)
    db.session.commit()
    flash('Projeto excluído com sucesso!', 'success')
    return redirect(url_for('projects.list_projects'))


@projects_bp.route('/projects/export-excel')
@login_required
@tenant_required
def export_excel():
    """Export all projects to Excel with multiple sheets."""
    if not EXCEL_AVAILABLE:
        flash('Funcionalidade de exportação Excel não disponível. Instale openpyxl.', 'danger')
        return redirect(url_for('projects.list_projects'))

    tenant_id = get_current_tenant_id()

    # Filter all queries by tenant
    if tenant_id:
        projects = Project.query.filter_by(tenant_id=tenant_id).order_by(Project.created_at.desc()).all()
    else:
        projects = Project.query.order_by(Project.created_at.desc()).all()

    wb = Workbook()

    # Style definitions
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Sheet 1: Projects (Dados Gerais)
    ws_projects = wb.active
    ws_projects.title = 'Projetos'
    project_headers = ['ID', 'Código', 'Título', 'Descrição', 'Status', 'Categoria',
                       'Data Início', 'Data Fim', 'Orçamento (R$)', 'Fonte de Recursos',
                       'Responsável', 'Criado em', 'Atualizado em']
    _write_headers(ws_projects, project_headers, header_font, header_fill, header_alignment, thin_border)

    for row_idx, project in enumerate(projects, start=2):
        ws_projects.cell(row=row_idx, column=1, value=project.id)
        ws_projects.cell(row=row_idx, column=2, value=project.code)
        ws_projects.cell(row=row_idx, column=3, value=project.title)
        ws_projects.cell(row=row_idx, column=4, value=project.description or '')
        ws_projects.cell(row=row_idx, column=5, value=project.status)
        ws_projects.cell(row=row_idx, column=6, value=project.category or '')
        ws_projects.cell(row=row_idx, column=7, value=project.start_date.strftime('%d/%m/%Y') if project.start_date else '')
        ws_projects.cell(row=row_idx, column=8, value=project.end_date.strftime('%d/%m/%Y') if project.end_date else '')
        ws_projects.cell(row=row_idx, column=9, value=project.budget)
        ws_projects.cell(row=row_idx, column=10, value=project.funding_source or '')
        ws_projects.cell(row=row_idx, column=11, value=project.responsible.full_name if project.responsible else '')
        ws_projects.cell(row=row_idx, column=12, value=project.created_at.strftime('%d/%m/%Y %H:%M') if project.created_at else '')
        ws_projects.cell(row=row_idx, column=13, value=project.updated_at.strftime('%d/%m/%Y %H:%M') if project.updated_at else '')

    _auto_adjust_columns(ws_projects)

    # Sheet 2: Expenses (Despesas)
    ws_expenses = wb.create_sheet('Despesas')
    expense_headers = ['Projeto', 'Código Projeto', 'Descrição', 'Categoria', 'Valor (R$)',
                       'Data', 'Nº Recibo', 'Fornecedor', 'Status', 'Notas', 'Criado por', 'Criado em']
    _write_headers(ws_expenses, expense_headers, header_font, header_fill, header_alignment, thin_border)

    if tenant_id:
        expenses = Expense.query.filter_by(tenant_id=tenant_id).join(Project).order_by(Project.code, Expense.date.desc()).all()
    else:
        expenses = Expense.query.join(Project).order_by(Project.code, Expense.date.desc()).all()

    for row_idx, expense in enumerate(expenses, start=2):
        ws_expenses.cell(row=row_idx, column=1, value=expense.project.title)
        ws_expenses.cell(row=row_idx, column=2, value=expense.project.code)
        ws_expenses.cell(row=row_idx, column=3, value=expense.description)
        ws_expenses.cell(row=row_idx, column=4, value=expense.category)
        ws_expenses.cell(row=row_idx, column=5, value=expense.amount)
        ws_expenses.cell(row=row_idx, column=6, value=expense.date.strftime('%d/%m/%Y') if expense.date else '')
        ws_expenses.cell(row=row_idx, column=7, value=expense.receipt_number or '')
        ws_expenses.cell(row=row_idx, column=8, value=expense.supplier or '')
        ws_expenses.cell(row=row_idx, column=9, value=expense.status)
        ws_expenses.cell(row=row_idx, column=10, value=expense.notes or '')
        ws_expenses.cell(row=row_idx, column=11, value=expense.created_by.full_name if expense.created_by else '')
        ws_expenses.cell(row=row_idx, column=12, value=expense.created_at.strftime('%d/%m/%Y %H:%M') if expense.created_at else '')

    _auto_adjust_columns(ws_expenses)

    # Sheet 3: Resources (Recursos)
    ws_resources = wb.create_sheet('Recursos')
    resource_headers = ['Projeto', 'Código Projeto', 'Tipo', 'Nome', 'Função',
                        'Horas Alocadas', 'Custo/Hora (R$)', 'Data Início', 'Data Fim', 'Status', 'Notas']
    _write_headers(ws_resources, resource_headers, header_font, header_fill, header_alignment, thin_border)

    if tenant_id:
        resources = Resource.query.filter_by(tenant_id=tenant_id).join(Project).order_by(Project.code).all()
    else:
        resources = Resource.query.join(Project).order_by(Project.code).all()

    for row_idx, resource in enumerate(resources, start=2):
        ws_resources.cell(row=row_idx, column=1, value=resource.project.title)
        ws_resources.cell(row=row_idx, column=2, value=resource.project.code)
        ws_resources.cell(row=row_idx, column=3, value=resource.type)
        ws_resources.cell(row=row_idx, column=4, value=resource.name)
        ws_resources.cell(row=row_idx, column=5, value=resource.role or '')
        ws_resources.cell(row=row_idx, column=6, value=resource.hours_allocated)
        ws_resources.cell(row=row_idx, column=7, value=resource.hourly_cost)
        ws_resources.cell(row=row_idx, column=8, value=resource.start_date.strftime('%d/%m/%Y') if resource.start_date else '')
        ws_resources.cell(row=row_idx, column=9, value=resource.end_date.strftime('%d/%m/%Y') if resource.end_date else '')
        ws_resources.cell(row=row_idx, column=10, value=resource.status)
        ws_resources.cell(row=row_idx, column=11, value=resource.notes or '')

    _auto_adjust_columns(ws_resources)

    # Sheet 4: Milestones (Cronograma)
    ws_milestones = wb.create_sheet('Cronograma')
    milestone_headers = ['Projeto', 'Código Projeto', 'Marco', 'Descrição',
                         'Data Início', 'Data Fim', 'Progresso (%)', 'Status', 'Ordem']
    _write_headers(ws_milestones, milestone_headers, header_font, header_fill, header_alignment, thin_border)

    if tenant_id:
        milestones = Milestone.query.filter_by(tenant_id=tenant_id).join(Project).order_by(Project.code, Milestone.order).all()
    else:
        milestones = Milestone.query.join(Project).order_by(Project.code, Milestone.order).all()

    for row_idx, milestone in enumerate(milestones, start=2):
        ws_milestones.cell(row=row_idx, column=1, value=milestone.project.title)
        ws_milestones.cell(row=row_idx, column=2, value=milestone.project.code)
        ws_milestones.cell(row=row_idx, column=3, value=milestone.title)
        ws_milestones.cell(row=row_idx, column=4, value=milestone.description or '')
        ws_milestones.cell(row=row_idx, column=5, value=milestone.start_date.strftime('%d/%m/%Y') if milestone.start_date else '')
        ws_milestones.cell(row=row_idx, column=6, value=milestone.end_date.strftime('%d/%m/%Y') if milestone.end_date else '')
        ws_milestones.cell(row=row_idx, column=7, value=milestone.progress)
        ws_milestones.cell(row=row_idx, column=8, value=milestone.status)
        ws_milestones.cell(row=row_idx, column=9, value=milestone.order)

    _auto_adjust_columns(ws_milestones)

    # Sheet 5: Timesheets (Apontamento de Horas)
    ws_timesheets = wb.create_sheet('Apontamento de Horas')
    timesheet_headers = ['Projeto', 'Código Projeto', 'Usuário', 'Data', 'Horas',
                         'Atividade', 'Marco', 'Recurso', 'Notas']
    _write_headers(ws_timesheets, timesheet_headers, header_font, header_fill, header_alignment, thin_border)

    if tenant_id:
        timesheets = Timesheet.query.filter_by(tenant_id=tenant_id).join(Project).order_by(Project.code, Timesheet.date.desc()).all()
    else:
        timesheets = Timesheet.query.join(Project).order_by(Project.code, Timesheet.date.desc()).all()

    for row_idx, ts in enumerate(timesheets, start=2):
        ws_timesheets.cell(row=row_idx, column=1, value=ts.project.title)
        ws_timesheets.cell(row=row_idx, column=2, value=ts.project.code)
        ws_timesheets.cell(row=row_idx, column=3, value=ts.user.full_name if ts.user else '')
        ws_timesheets.cell(row=row_idx, column=4, value=ts.date.strftime('%d/%m/%Y') if ts.date else '')
        ws_timesheets.cell(row=row_idx, column=5, value=ts.hours)
        ws_timesheets.cell(row=row_idx, column=6, value=ts.activity)
        ws_timesheets.cell(row=row_idx, column=7, value=ts.milestone.title if ts.milestone else '')
        ws_timesheets.cell(row=row_idx, column=8, value=ts.resource.name if ts.resource else '')
        ws_timesheets.cell(row=row_idx, column=9, value=ts.notes or '')

    _auto_adjust_columns(ws_timesheets)

    # Sheet 6: Project-Call Links (Chamadas Públicas Vinculadas)
    ws_calls = wb.create_sheet('Chamadas Vinculadas')
    call_headers = ['Projeto', 'Código Projeto', 'Chamada', 'Fonte', 'Tema',
                    'Data Vinculação', 'Status Vínculo', 'Notas']
    _write_headers(ws_calls, call_headers, header_font, header_fill, header_alignment, thin_border)

    if tenant_id:
        project_calls = ProjectCall.query.filter_by(tenant_id=tenant_id).join(Project).order_by(Project.code).all()
    else:
        project_calls = ProjectCall.query.join(Project).order_by(Project.code).all()

    for row_idx, pc in enumerate(project_calls, start=2):
        ws_calls.cell(row=row_idx, column=1, value=pc.project.title)
        ws_calls.cell(row=row_idx, column=2, value=pc.project.code)
        ws_calls.cell(row=row_idx, column=3, value=pc.public_call.title if pc.public_call else '')
        ws_calls.cell(row=row_idx, column=4, value=pc.public_call.source if pc.public_call else '')
        ws_calls.cell(row=row_idx, column=5, value=pc.public_call.theme if pc.public_call else '')
        ws_calls.cell(row=row_idx, column=6, value=pc.linked_at.strftime('%d/%m/%Y') if pc.linked_at else '')
        ws_calls.cell(row=row_idx, column=7, value=pc.status)
        ws_calls.cell(row=row_idx, column=8, value=pc.notes or '')

    _auto_adjust_columns(ws_calls)

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    tenant_name = current_user.tenant.slug if current_user.tenant else 'all'
    filename = f'projetos_pd_{tenant_name}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


def _write_headers(ws, headers, font, fill, alignment, border):
    """Write headers to a worksheet with styling."""
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = font
        cell.fill = fill
        cell.alignment = alignment
        cell.border = border


def _auto_adjust_columns(ws):
    """Auto-adjust column widths based on content."""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except (TypeError, AttributeError):
                pass
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
        ws.column_dimensions[column_letter].width = adjusted_width
